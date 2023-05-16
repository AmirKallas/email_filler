import openpyxl
import xlsxwriter
import datetime
from utilities import Utilities

HRR_PATH = 'C:\\Users\\amirk\\Desktop\\filler_data\\DBCLabHRR.xlsx'
TRANSFERT_PATH = 'C:\\Users\\amirk\\Desktop\\filler_data\\DBCLabTransfer.xlsx'

hrrWorkbook = openpyxl.load_workbook(HRR_PATH, read_only=True, data_only=True)
#transfertWorkbook = openpyxl.load_workbook(TRANSFERT_PATH, read_only=True, data_only=True)


candidatiHrr = hrrWorkbook['Candidati']
anagSkill = hrrWorkbook['AnagSkill']

# with open(HRR_PATH, 'r') as 

TODAY  = str(datetime.date.today())
stato = 'CV al Cliente'


rows = candidatiHrr.max_row
cols = candidatiHrr.max_column

rows_candidati = [] #lista di righe utili
rows_anagskill = []
for row_num in range(1, 5):
    if str(candidatiHrr.cell(row=row_num, column=21).value)[0:10] == TODAY and str(candidatiHrr.cell(row=row_num, column=20).value) == stato: #compara data odierna con data ria
        tempList = []
        for col_num in range(1, cols+1):
            value = candidatiHrr.cell(row=row_num, column=col_num).value
            tempList.append(str(value))
        rows_candidati.append(tempList)

# print(', '.join(row_cells))
# print(row_cells)

for candidato in rows_candidati:
    idCandidato = int(candidato[7])
    for row_num in range(1, 5):
        if anagSkill.cell(row=row_num, column=2).value == idCandidato:
            print('gino')
            tempList1 = []
            for col_num in range(1, cols+1):
                valore = candidatiHrr.cell(row=row_num, column=col_num).value
                tempList1.append(str(valore))
            rows_anagskill.append(tempList1)
            print(rows_anagskill)
            

transfertWorkbook = openpyxl.load_workbook(filename = TRANSFERT_PATH, read_only=True, data_only=True)
print('---------------------CICICICICIC----------------------------------------------------------')
# candidatiTrf = transfertWorkbook['Candidati']
# anagSkillTrf = transfertWorkbook['AnagSkill']

newWorkbook = xlsxwriter.Workbook('provaTransfer.xlsx')
print(newWorkbook)
for sheet_name in transfertWorkbook.sheetnames:
    print(sheet_name)
    sheet = transfertWorkbook[sheet_name]
    print(sheet)
    new_sheet = newWorkbook.add_worksheet(sheet_name)
    print(new_sheet)
    for row_num, row in enumerate(sheet.iter_rows()):
        for col_num, cell in enumerate(row):
            new_sheet.write(row_num, col_num, cell.value)
            
newWorkbook.close()
# print('---------------------CICICICICIC----------------------------------------------------------')


# al = Utilities.alpha
# for anag in rows_anagskill:
#     idxAl = 0
#     for elem in anag:
#         colIdx = 0
#         idx = al[idxAl]+str(colIdx)
#         anagSkillTrf
#         colIdx += 1
#     idxAl += 1    
# transfertWorkbook.save(TRANSFERT_PATH)
        
        
    
            

            
            

            
            

        
    




